import datetime
from datetime import date
from datetime import datetime
from django.shortcuts import render,redirect, get_object_or_404
from django.views import View
from django.http import HttpResponse
import pandas as pd
from plotly.offline import plot
import plotly.express as px
from django.contrib import messages
from django.core.mail import send_mail
from django.views.generic.base import TemplateView
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Fill, Side

from app import *
from .models import *
from .forms import *
from usuario.models import Tipo_Usuario
from fonoaudiologo.models import Fonoaudiologo


#index de la pagina principal con la fecha actual
def index(request):

    dia = date.today().day
    mes = date.today().month
    
    if mes == 1:
        mes = 'Ene'
        
    if mes == 2:
        mes = 'Feb'
        
    if mes == 3:
        mes = 'Mar'
        
    if mes == 4:
        mes = 'Abr'
        
    if mes == 5:
        mes = 'May'
        
    if mes == 6:
        mes = 'Jun'
        
    if mes == 7:
        mes = 'Jul'
        
    if mes == 8:
        mes = 'Ago'
        
    if mes == 9:
        mes = 'Sep'
        
    if mes == 10:
        mes = 'Oct'
        
    if mes == 11:
        mes = 'Nov'
        
    if mes == 12:
        mes = 'Dic'
       
    return render(request, 'index.html',{'mes':mes,'dia':dia})

#vista login de usuario
def login_vista(request):
    return render(request, 'login.html')

#pos login, dependiendo del usuario logueado se muestra una vista diferente
def post_login(request):
    tipo_usuarios=Tipo_Usuario.objects.all().filter(id_tipo_usuario=1)
    pacientes=Paciente.objects.all()

    if request.user.tipo_usuario_id==1:
        return render(request, 'pacientes/index_login.html',{'tipo_usuarios':tipo_usuarios,'pacientes':pacientes})
    elif request.user.tipo_usuario_id ==2:
        return render(request, 'neurologo/neurologo.html',{'tipo_usuarios':tipo_usuarios,'pacientes':pacientes})
    elif request.user.tipo_usuario_id ==3:
        return render(request, 'familiar/familiar.html',{'tipo_usuarios':tipo_usuarios,'pacientes':pacientes})
    elif request.user.tipo_usuario_id ==4:
        return render(request, 'fonoaudiologo/fonoaudiologo.html',{'tipo_usuarios':tipo_usuarios,'pacientes':pacientes})
    elif request.user.tipo_usuario_id ==5:
        return render(request, 'enfermera/enfermera.html',{'tipo_usuarios':tipo_usuarios,'pacientes':pacientes})
    elif request.user.is_staff ==1:
        return render(request, 'admin_paciente/admin_paciente.html',{'tipo_usuarios':tipo_usuarios,'pacientes':pacientes})
    elif request.user.tipo_usuario.nombre_tipo_usuario == "adminEnfermera":
        return render(request, 'enfermera/enfermera.html',{'tipo_usuarios':tipo_usuarios,'pacientes':pacientes})   
    else: 
        print("holaaaaaaaaaaaaa")
        return render(request, 'error_usuario.html',{'tipo_usuarios':tipo_usuarios,'pacientes':pacientes})

#vista del index de enfermera 
def enfermera(request):
    tipo_usuarios=Tipo_Usuario.objects.all().filter(id_tipo_usuario=request.user.tipo_usuario_id)
    enfermeras=Enfermera.objects.all
    return render(request, 'enfermera/enfermera.html',{'tipo_usuarios':tipo_usuarios,'enfermeras':enfermeras})

#vista de los pacientes de la enfermera logueada
def enfermera_paciente(request):
    enfermera_pacientes=Enfermera_paciente.objects.all().filter(username_enfermera=request.user.id)
    return render(request, 'enfermera/enfermera_paciente.html',{'enfermera_pacientes':enfermera_pacientes})

#informacion del paciente seleccionado por la enferemera
def enfermera_info_paciente(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente=username_paciente_id)
    neurologo_pacientes=Neurologo_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    familiar_pacientes=Familiar_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    receta=Receta.objects.all().filter(username_paciente_id=username_paciente_id)
    bitacoras =Bitacora.objects.all().filter(username_paciente_id=username_paciente_id)
    paciente_documento=Paciente_Documento.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'enfermera/enfermera_info_paciente.html',{'paciente_documento':paciente_documento,'bitacoras':bitacoras,'receta':receta,'familiar_pacientes':familiar_pacientes,'pacientes':pacientes,'neurologo_pacientes':neurologo_pacientes})

#informacion del paciente seleccionado por la enferemera desde el admin
def enfermera_vista_info_paciente(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente=username_paciente_id)
    neurologo_pacientes=Neurologo_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    familiar_pacientes=Familiar_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    receta=Receta.objects.all().filter(username_paciente_id=username_paciente_id)
    bitacoras =Bitacora.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_enfermera/vista_enfermera.html',{'bitacoras':bitacoras,'receta':receta,'familiar_pacientes':familiar_pacientes,'pacientes':pacientes,'neurologo_pacientes':neurologo_pacientes})

#vista del index de neurologo 
def neurologo(request):
    tipo_usuarios=Tipo_Usuario.objects.all().filter(id_tipo_usuario=request.user.tipo_usuario_id)
    neurologos=Neurologo.objects.all().filter(id_neurologo=1)
    return render(request, 'neurologo/neurologo.html',{'tipo_usuarios':tipo_usuarios,'neurologos':neurologos})

#vista de los pacientes del neurologo logueado
def neurologo_paciente(request):
    neurologo_pacientes=Neurologo_paciente.objects.all().filter(username_neurologo=request.user.id)
    return render(request, 'neurologo/neurologo_paciente.html',{'neurologo_pacientes':neurologo_pacientes})

#informacion del paciente seleccionado por el neurologo
def neurologo_info_paciente(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente=username_paciente_id)
    neurologo_pacientes=Neurologo_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    familiar_pacientes=Familiar_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    receta=Receta.objects.all().filter(username_paciente_id=username_paciente_id)
    estado_animos = Estado_animo.objects.all().filter(username_paciente_id=username_paciente_id)
    automonitoreo=Automonitoreo.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'neurologo/neurologo_info_paciente.html',{'automonitoreo':automonitoreo,'estado_animos':estado_animos,'receta':receta,'familiar_pacientes':familiar_pacientes,'pacientes':pacientes,'neurologo_pacientes':neurologo_pacientes})


#vista del index de fonoaudiologo 
def fonoaudiologo(request):
    tipo_usuarios=Tipo_Usuario.objects.all().filter(id_tipo_usuario=request.user.tipo_usuario_id)
    fonoaudiologos=Fonoaudiologo.objects.all().filter(id_fonoaudiologo=1)
    return render(request, 'fonoaudiologo/fonoaudiologo.html',{'tipo_usuarios':tipo_usuarios,'fonoaudiologo':fonoaudiologos})



#informacion del paciente seleccionado por el fonoaudiologo desde admin
def fonoaudiologo_vista_info_paciente(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente=username_paciente_id)
    fonoaudiologo_pacientes=Fonoaudiologo_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    familiar_pacientes=Familiar_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    receta=Receta.objects.all().filter(username_paciente_id=username_paciente_id)
    estado_animos = Estado_animo.objects.all().filter(username_paciente_id=username_paciente_id)
    automonitoreo=Automonitoreo.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/vista_fonoaudiologo.html',{'automonitoreo':automonitoreo,'estado_animos':estado_animos,'receta':receta,'familiar_pacientes':familiar_pacientes,'pacientes':pacientes,'fonoaudiologo_pacientes':fonoaudiologo_pacientes})



#recetas del paciente
def receta(request):
    pacientes=Paciente.objects.all().filter(username_paciente=request.user.id)
    recetas=Receta.objects.all().filter(username_paciente=request.user.id)

    return render(request, 'pacientes/receta.html',{'recetas':recetas,'pacientes':pacientes})

#vista del index de familiar 
def familiar(request):
    tipo_usuarios=Tipo_Usuario.objects.all().filter(id_tipo_usuario= request.user.tipo_usuario_id)
    familiares=Familiar.objects.all().filter(id_familiar=2)
    return render(request, 'familiar/familiar.html',{'familiares':familiares,'tipo_usuarios':tipo_usuarios})

#vista de los pacientes del familiar logueado
def familiar_paciente(request):
    familiar_pacientes=Familiar_paciente.objects.all().filter(username_familiar=request.user.id)
    return render(request, 'familiar/familiar_paciente.html',{'familiar_pacientes':familiar_pacientes})

#informacion del paciente seleccionado por el fonoaudiologo
def familiar_info_paciente(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente=username_paciente_id)
    familiar_pacientes=Familiar_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'familiar/familiar_info_paciente.html',{'pacientes':pacientes,'familiar_pacientes':familiar_pacientes})

#vista de los pacientes en el sistema visto por admin
def admin_paciente(request):
    pacientes=Paciente.objects.all()
    return render(request, 'admin_paciente/admin_paciente.html',{'pacientes':pacientes})

#informacion del paciente seleccionado en el admin
def admin_receta_paciente(request,username_paciente_id):
    recetas=Receta.objects.all().filter(username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente=username_paciente_id)
    enfermera_pacientes=Enfermera_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    neurologo_pacientes=Neurologo_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    familiar_pacientes=Familiar_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    fonoaudiologo_pacientes=Fonoaudiologo_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/admin_receta_paciente.html',{'fonoaudiologo_pacientes':fonoaudiologo_pacientes,
    'recetas':recetas,'enfermera_pacientes':enfermera_pacientes,'pacientes':pacientes,'neurologo_pacientes':neurologo_pacientes,'familiar_pacientes':familiar_pacientes})

#vista de los familiares en el sistema visto por admin
def admin_familiar(request):
    familiares=Familiar.objects.all()
    return render(request, 'admin_familiar/admin_familiar.html',{'familiares':familiares})

#informacion del familiar seleccionado en el admin
def admin_info_familiar(request,username_familiar_id):
    familiares=Familiar.objects.all().filter(username_familiar=username_familiar_id)
    familiar_pacientes=Familiar_paciente.objects.all().filter(username_familiar_id=username_familiar_id)
    return render(request, 'admin_familiar/admin_info_familiar.html',{'familiares':familiares,'familiar_pacientes':familiar_pacientes})

#vista de las enfermeras en el sistema visto por admin
def admin_enfermera(request):
    enfermeras=Enfermera.objects.all()
    return render(request, 'admin_enfermera/admin_enfermera.html',{'enfermeras':enfermeras})

#informacion de las enfermeras seleccionado en el admin
def admin_info_enfermera(request,username_enfermera_id):
    enfermeras=Enfermera.objects.all().filter(username_enfermera=username_enfermera_id)
    enfermera_pacientes=Enfermera_paciente.objects.all().filter(username_enfermera=username_enfermera_id)
    return render(request, 'admin_enfermera/admin_info_enfermera.html',{'enfermeras':enfermeras,'enfermera_pacientes':enfermera_pacientes})

#vista de los neurologos en el sistema visto por admin
def admin_neurologo(request):
    neurologos=Neurologo.objects.all()
    return render(request, 'admin_neurologo/admin_neurologo.html',{'neurologos':neurologos})

#informacion del neurologo seleccionado en el admin
def admin_info_neurologo(request,username_neurologo_id):
    neurologos=Neurologo.objects.all().filter(username_neurologo=username_neurologo_id)
    neurologo_pacientes=Neurologo_paciente.objects.all().filter(username_neurologo_id=username_neurologo_id)
    return render(request, 'admin_neurologo/admin_info_neurologo.html',{'neurologos':neurologos,'neurologo_pacientes':neurologo_pacientes})

#vista de los fonoaudiologos en el sistema visto por admin
def admin_fonoaudiologo(request):
    fonoaudiologos=Fonoaudiologo.objects.all()
    return render(request, 'admin_fonoaudiologo/admin_fonoaudiologo.html',{'fonoaudiologos':fonoaudiologos})

#informacion del fonoaudiologo seleccionado en el admin
def admin_info_fonoaudiologo(request,username_fonoaudiologo_id):
    fonoaudiologos=Fonoaudiologo.objects.all().filter(username_fonoaudiologo=username_fonoaudiologo_id)
    fonoaudiologo_pacientes=Fonoaudiologo_paciente.objects.all().filter(username_fonoaudiologo_id=username_fonoaudiologo_id)
    return render(request, 'admin_fonoaudiologo/admin_info_fonoaudiologo.html',{'fonoaudiologos':fonoaudiologos,'fonoaudiologo_pacientes':fonoaudiologo_pacientes})





#vinculacion enfermera_paciente
def enf_pac(request,id):
    preregistros=Preregistro.objects.all().filter(id=id)
    current_user =get_object_or_404(Enfermera, pk=request.user.pk)
    
    if request.method == 'POST':
        form = FormEnfPac(request.POST)
        if form.is_valid():
            estado_emocion = form.save(commit=False)
            estado_emocion.username_enfermera = current_user
            estado_emocion.save()
        return redirect('neu_pac',id)
    else:
        form=FormEnfPac()
    return render(request, 'enfermera/vincular_paciente.html', {'form':form, 'preregistros':preregistros})

#vinculacion neurologo_paciente
def neu_pac(request,id):
    preregistros=Preregistro.objects.all().filter(id=id)
    current_user =get_object_or_404(Preregistro, id=id)
    if request.method == 'POST':
        form = FormNeuPac(request.POST)
        if form.is_valid():
            estado_emocion = form.save(commit=False)
            estado_emocion.username_neurologo = current_user.neurologo
            estado_emocion.save()
        return redirect('enfermera_paciente')
    else:
        form=FormNeuPac()
    return render(request, 'enfermera/vincular_paciente.html', {'form':form, 'preregistros':preregistros})


    def post(self,request,id):
        id_paciente =request.POST['id']
        username_paciente=request.POST['username_paciente']
        rut_paciente=request.POST['rut_paciente']
        nombre_paciente=request.POST['nombre_paciente']
        apellido_paciente=request.POST['apellido_paciente']
        diabetes=request.POST['diabetes']
        hipertension=request.POST['hipertension']
        direccion_paciente=request.POST['direccion_paciente']
        comuna=request.POST['comuna']
        email_paciente=request.POST['email_paciente']
        telefono_paciente=request.POST['telefono_paciente']
        whatsaap_paciente=request.POST['whatsaap_paciente']
        celular_paciente=request.POST['celular_paciente']
        telegram_paciente=request.POST['telegram_paciente']
        new_paciente=Paciente.objects.create(id_paciente=id_paciente,
                                            username_paciente_id=username_paciente,
                                            rut_paciente=rut_paciente,
                                            nombre_paciente=nombre_paciente,
                                            apellido_paciente=apellido_paciente,
                                            diabetes_id=diabetes,
                                            hipertension_id=hipertension,
                                            direccion_paciente=direccion_paciente,
                                            comuna_id=comuna,
                                            email_paciente=email_paciente,
                                            telefono_paciente=telefono_paciente,
                                            whatsaap_paciente=whatsaap_paciente,
                                            celular_paciente=celular_paciente,
                                            telegram_paciente=telegram_paciente
                                            )
        new_paciente.save()
        return redirect("enf_pac", id)

#formulario del estado de animo
def estado_animo_form(request):
    current_user =get_object_or_404(Paciente, pk=request.user.pk)
    estado_animos = Estado_animo.objects.all().filter(username_paciente_id=request.user.id)
    if request.method == 'POST':
        form = FormEmociones(request.POST)
        if form.is_valid():
            estado_emocion = form.save(commit=False)
            estado_emocion.username_paciente = current_user
            estado_emocion.save()
        return redirect('estado_animo_form')
    else:
        form=FormEmociones()
    return render(request, 'pacientes/forms/estado_animo_form.html', {'form':form,'estado_animos':estado_animos})

#crear id de telegram
def telegram(request, id):
    paciente = get_object_or_404(Paciente, username_paciente = id)
    data = {
        'form': FormTelegram(instance=paciente)
    }
    if request.method == 'POST':
        formulario = FormTelegram(data= request.POST, instance=paciente)
        if formulario.is_valid():
            formulario.save()
            return redirect(to="post_login")
        else:
            data["form"] = formulario  
    return render(request, 'pacientes/telegram.html', data)

#formulario de automonitoreo
def automonitoreo_form(request):
    current_user =get_object_or_404(Paciente, pk=request.user.pk)
    automonitoreo=Automonitoreo.objects.all().filter(username_paciente_id=request.user.id)
    if request.method == 'POST':
        form = FormAutomonitoreo(request.POST)
        if form.is_valid():
            estado_emocion = form.save(commit=False)
            estado_emocion.username_paciente = current_user
            estado_emocion.save()
        return redirect('automonitoreo_form')
    else:
        form=FormAutomonitoreo()
    return render(request, 'pacientes/forms/automonitoreo_form.html', {'form':form,'automonitoreo':automonitoreo})

#grafico del animo 
def chart_animo(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    qs = Estado_animo.objects.all().filter(username_paciente_id=username_paciente_id)
    animo_data=[{
        'Animo': x.animo.nombre_animo,
        'Start': x.fecha_estado_animo.astimezone() ,
        'Finish': x.fecha_estado_animo.astimezone() + datetime.timedelta(seconds=3600),
        'Estado Animo': x.animo.nombre_animo}
        for x in qs
]
    df = pd.DataFrame(animo_data)
    fig = px.timeline(
        df, x_start="Start",x_end="Finish",y="Animo", color="Estado Animo"
    )
    print(df)
    fig.update_yaxes(autorange="reversed")
    gantt_plot = plot(fig, output_type="div")
    context={'plot_div':gantt_plot, 'pacientes':pacientes}
    return render(request, 'neurologo/chart_animo.html', context)

#grafico del automonitoreo
def chart_automonitoreo(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    qs = Automonitoreo.objects.all().filter(username_paciente_id=username_paciente_id)
    automonitoreo_data=[{
        'Autmonitoreo': x.estado_monitoreo.estado_monitoreo,
        'Start': x.fecha_automonitoreo.astimezone() ,
        'Finish': x.fecha_automonitoreo.astimezone() + datetime.timedelta(seconds=3600),
        'Estado Automonitoreo': x.estado_monitoreo.estado_monitoreo}
        for x in qs
]
    df = pd.DataFrame(automonitoreo_data)
    fig = px.timeline(
        df, x_start="Start",x_end="Finish",y="Autmonitoreo", color="Estado Automonitoreo"
    )
    print(df)
    fig.update_yaxes(autorange="reversed")
    gantt_plot = plot(fig, output_type="div")
    context={'plot_div':gantt_plot, 'pacientes':pacientes}
    return render(request, 'neurologo/chart_automonitoreo.html', context)

#audios del paciente logueado
def audio(request,username_paciente_id):
    audios=Audio.objects.all().filter(username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'pacientes/audio.html',{'audios':audios,'pacientes':pacientes})





#vocalizaciones 
def vocalizacion_paciente(request,username_paciente_id):
    vocalizacion=Vocalizacion.objects.all().filter(username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'pacientes/vocalizacion.html',{'vocalizacion':vocalizacion,'pacientes':pacientes})



#intensidades
def intensidad_paciente(request,username_paciente_id):
    intensidad=Intensidad.objects.all().filter(username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'pacientes/intensidad.html',{'intensidad':intensidad,'pacientes':pacientes})

def terapias_fono(request):
    return render(request, 'pacientes/terapias_fono.html')

#audios del paciente desde admin
def vista_audio_paciente(request,username_paciente_id):
    audios=Audio.objects.all().filter(username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/audio_paciente.html',{'audios':audios,'pacientes':pacientes})

#vocalizaciones desde vista admin
def vista_vocalizacion(request,username_paciente_id):
    vocalizacion=Vocalizacion.objects.all().filter(username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/vocalizacion.html',{'vocalizacion':vocalizacion,'pacientes':pacientes})

#intensidades desde vista admin
def vista_intensidad(request,username_paciente_id):
    intensidad=Intensidad.objects.all().filter(username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/intensidad.html',{'intensidad':intensidad,'pacientes':pacientes})

#editar comentario de voalizaciones
def editar_comentario(request, id):
    vocalizacion = Vocalizacion.objects.get(id=id)
    vocalizaciones = Vocalizacion.objects.all().filter(id=id)
    formulario = FormComentario(request.POST or None, request.FILES or None, instance=vocalizacion)
    if formulario.is_valid() and request.POST:
        formulario.save()
        # messages.success(request, 'Producto editado')
        return redirect('vocalizacion', username_paciente_id=vocalizacion.username_paciente_id)
    return render(request, 'fonoaudiologo/editar_comentario.html', {'formulario': formulario,'vocalizaciones': vocalizaciones})

#editar comentario de intensidad
def editar_comentario_intensidad(request, id):
    intensidad = Intensidad.objects.get(id=id)
    intensidades = Intensidad.objects.all().filter(id=id)
    formulario = FormComentarioIntensidad(request.POST or None, request.FILES or None, instance=intensidad)
    if formulario.is_valid() and request.POST:
        formulario.save()
        return redirect('intensidad', username_paciente_id=intensidad.username_paciente_id)
    return render(request, 'fonoaudiologo/editar_comentario_intensidad.html', {'formulario': formulario,'intensidades': intensidades})

#editar comentario de bitacora
def editar_comentario_bitacora(request, id):
    bitacora = Bitacora.objects.get(id=id)
    bitacoras = Bitacora.objects.all().filter(id=id)
    formulario = FormBitacora(request.POST or None,  instance=bitacora)
    if formulario.is_valid() and request.POST:
        formulario.save()
        return redirect('enfermera_info_paciente', username_paciente_id=bitacora.username_paciente_id)
    return render(request, 'enfermera/form_editar_bitacora.html', {'formulario': formulario, 'bitacoras': bitacoras})

#editar medicamento de la receta
def editar_receta(request, id_receta):
    receta = Receta.objects.get(id_receta=id_receta)
    recetas = Receta.objects.all().filter(id_receta=id_receta)
    formulario = FormReceta(request.POST or None,  instance=receta)
    if formulario.is_valid() and request.POST:
        formulario.save()
        return redirect('enfermera_info_paciente', username_paciente_id=receta.username_paciente_id)
    return render(request, 'enfermera/form_editar_receta.html', {'formulario': formulario, 'recetas': recetas})

#eliminar receta
def eliminar_receta(request,id_receta):
    receta = Receta.objects.get(id_receta=id_receta)
    receta.delete()
    return redirect(to="enfermera_info_paciente", username_paciente_id=receta.username_paciente_id)

#crear receta 
def agregar_receta(request,username_paciente_id):
    current_user =get_object_or_404(Enfermera, pk=request.user.pk)
    current_pac = get_object_or_404(Paciente, username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    if request.method == 'POST':
        form = FormReceta(request.POST)
        if form.is_valid():
            estado_emocion = form.save(commit=False)
            estado_emocion.username_paciente = current_pac
            estado_emocion.username_enfermera = current_user
            estado_emocion.save()
        return redirect('enfermera_info_paciente', current_pac.username_paciente_id )
    else:
        form=FormReceta()
    return render(request, 'enfermera/form_receta.html', {'form':form,'pacientes':pacientes})






#informacion del paciente vista como neurologo desde admin
def neurologo_info_paciente_vista(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente=username_paciente_id)
    neurologo_pacientes=Neurologo_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    familiar_pacientes=Familiar_paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    receta=Receta.objects.all().filter(username_paciente_id=username_paciente_id)
    estado_animos = Estado_animo.objects.all().filter(username_paciente_id=username_paciente_id)
    automonitoreo=Automonitoreo.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_neurologo/vista_neurologo.html',{'automonitoreo':automonitoreo,'estado_animos':estado_animos,'receta':receta,'familiar_pacientes':familiar_pacientes,'pacientes':pacientes,'neurologo_pacientes':neurologo_pacientes})

#grafico animo desde admin
def chart_vista_animo(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    qs = Estado_animo.objects.all().filter(username_paciente_id=username_paciente_id)
    animo_data=[{
        'Animo': x.animo.nombre_animo,
        'Start': x.fecha_estado_animo.astimezone() ,
        'Finish': x.fecha_estado_animo.astimezone() + datetime.timedelta(seconds=3600),
        'Estado Animo': x.animo.nombre_animo}
        for x in qs
]
    df = pd.DataFrame(animo_data)
    fig = px.timeline(
        df, x_start="Start",x_end="Finish",y="Animo", color="Estado Animo"
    )
    print(df)
    fig.update_yaxes(autorange="reversed")
    gantt_plot = plot(fig, output_type="div")
    context={'plot_div':gantt_plot, 'pacientes':pacientes}
    return render(request, 'admin_paciente/vistas_neurologo/vista_grafico_animo.html', context)

#grafico automonitoreo desde admin
def chart_vista_automonitoreo(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    qs = Automonitoreo.objects.all().filter(username_paciente_id=username_paciente_id)
    automonitoreo_data=[{
        'Autmonitoreo': x.estado_monitoreo.estado_monitoreo,
        'Start': x.fecha_automonitoreo.astimezone() ,
        'Finish': x.fecha_automonitoreo.astimezone() + datetime.timedelta(seconds=3600),
        'Estado Automonitoreo': x.estado_monitoreo.estado_monitoreo}
        for x in qs
]
    df = pd.DataFrame(automonitoreo_data)
    fig = px.timeline(
        df, x_start="Start",x_end="Finish",y="Autmonitoreo", color="Estado Automonitoreo"
    )
    print(df)
    fig.update_yaxes(autorange="reversed")
    gantt_plot = plot(fig, output_type="div")
    context={'plot_div':gantt_plot, 'pacientes':pacientes}
    return render(request, 'admin_paciente/vistas_neurologo/vista_grafico_automonitoreo.html', context)

#links graficos admnin
def grafico_admin(request):
    return render(request, 'admin_grafico/graficos_admin.html')

#reporte excel de audios
class ReporteAudios(TemplateView):
    def get(self, request, *arg, **kwargs):
        query = Audio.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'REPORTE DE AUDIOS'
        ws.merge_cells('B1:P1')
        ws.title = 'Audios' 
        ws.row_dimensions[1].height = 20
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 11
        ws.column_dimensions['D'].width = 26

        ws.column_dimensions['E'].width = 6
        ws.column_dimensions['F'].width = 6
        ws.column_dimensions['G'].width = 11
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 11
        ws.column_dimensions['K'].width = 8
        ws.column_dimensions['L'].width = 8
        ws.column_dimensions['M'].width = 11
        ws.column_dimensions['N'].width = 9
        ws.column_dimensions['O'].width = 9
        ws.column_dimensions['P'].width = 11

        ws['B1'].alignment = Alignment(horizontal ="center", vertical = "center") 
        ws['B1'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['B1'].fill = PatternFill(start_color = '76933C', end_color = '76933C', fill_type = "solid")
        ws['B1'].font = Font(name = 'Calibri', size = 14, bold = True )
        
        #Titulos
        ws['B3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['B3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['B3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['B3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['B3'] = 'Nombre Paciente'

        ws['C3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['C3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['C3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['C3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['C3'] = 'Fecha de Envio'

        ws['D3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['D3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['D3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['D3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['D3'] = 'URL Audio'

        ws['E3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['E3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['E3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['E3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['E3'] = 'Jitter ppq5'

        ws['F3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['F3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['F3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['F3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['F3'] = 'Jitter rap'

        ws['G3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['G3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['G3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['G3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['G3'] = 'Maximum pitch'

        ws['H3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['H3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['H3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['H3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['H3'] = 'Error jitter ppq5'

        
        ws['I3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['I3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                               top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['I3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['I3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['I3'] = 'Error jitter rap'


        ws['J3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['J3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                              top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['J3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['J3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['J3'] = 'Error maximum pitch'







        ws['K3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['K3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['K3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['K3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['K3'] = 'Jitter ppq5 AI'

        ws['L3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['L3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['L3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['L3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['L3'] = 'Jitter rap AI'

        ws['M3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['M3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['M3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['M3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['M3'] = 'Maximum pitch AI'

        ws['N3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['N3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['N3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['N3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['N3'] = 'Error jitter ppq5 AI'

        
        ws['O3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['O3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                               top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['O3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['O3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['O3'] = 'Error jitter rap AI'


        ws['P3'].alignment = Alignment(horizontal ="center", vertical = "center", wrap_text=True) 
        ws['P3'].border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                              top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
        ws['P3'].fill = PatternFill(start_color = 'C4D79B', end_color = 'C4D79B', fill_type = "solid")
        ws['P3'].font = Font(name = 'Calibri', size = 12, bold = True )
        ws['P3'] = 'Error maximum pitch AI'
        cont = 4

        #Traer Informacion
        for q in query:
            ws.cell(row = cont, column = 2).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 2).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 2).font = Font(name = 'Calibri', size = 10)                         
            ws.cell(row = cont, column = 2).value =  str(q.username_paciente)

            ws.cell(row = cont, column = 3).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 3).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 3).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 3).value = str(q.timestamp.date())

            ws.cell(row = cont, column = 4).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 4).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 4).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 4).value = str(q.url_archivo_audio)
            # ws.cell(row = cont, column = 4).value = 'https://vozparkinson.pythonanywhere.com/' +str(q.url_archivo_audio)

            ws.cell(row = cont, column = 5).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 5).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 5).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 5).value = str(q.jitter_ppq5)

            ws.cell(row = cont, column = 6).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 6).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 6).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 6).value = str(q.jitter_rap)

            ws.cell(row = cont, column = 7).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 7).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 7).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 7).value = str(q.maximum_pitch)
            
            ws.cell(row = cont, column = 8).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 8).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 8).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 8).value = str(q.error_jitter_ppq5)


            ws.cell(row = cont, column = 9).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 9).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 9).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 9).value = str(q.error_jitter_rap)


            ws.cell(row = cont, column = 10).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 10).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 10).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 10).value = str(q.error_maximum_pitch)










            ws.cell(row = cont, column = 11).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 11).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 11).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 11).value = str(q.jitter_ppq5_IA)

            ws.cell(row = cont, column = 12).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 12).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 12).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 12).value = str(q.jitter_rap_IA)

            ws.cell(row = cont, column = 13).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 13).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 13).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 13).value = str(q.maximum_pitch_IA)
            
            ws.cell(row = cont, column = 14).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 14).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 14).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 14).value = str(q.error_jitter_ppq5_IA)


            ws.cell(row = cont, column = 15).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 15).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 15).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 15).value = str(q.error_jitter_rap_IA)


            ws.cell(row = cont, column = 16).alignment = Alignment(horizontal ="center", vertical = "center")
            ws.cell(row = cont, column = 16).border = Border(left = Side(border_style = "thin"), right = Side(border_style = "thin"),
                                    top = Side(border_style = "thin"), bottom = Side(border_style = "thin"))
            ws.cell(row = cont, column = 16).font = Font(name = 'Calibri', size = 10)
            ws.cell(row = cont, column = 16).value = str(q.error_maximum_pitch_IA)


            cont+=1
                
        #Establecer nombre archivo
        nombre_archivo = "InformeAudios.xlsx"
        #Definir tipo respuesta
        response = HttpResponse(content_type = "application/ms-excel")
        contenido = "attachment; filename = {0}".format(nombre_archivo)
        response["Content-Disposition"] = contenido
        wb.save(response)  
        return response

#grafico cantidad audios
def grafico_audios_admin(request):



    entradas = Audio.objects.all()
    salidas = Vocalizacion.objects.all()
    oportunidades = Intensidad.objects.all()

    dias_total = []
    for e in entradas:
        dias_total.append(e.timestamp.date())
        # print(dias_total)
        # print("list original", dias_total)
        convert_list_to_set = set(dias_total)
        # print("Set is: ",convert_list_to_set)
        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)
        dias_total = list(convert_list_to_set)
        # print("Removed duplicates from original list: ",dias_total)
        dias_total.sort()
        #crear diccionario para grafico
        diccionario = {}
        for i in dias_total:
                        diccionario[i] = 0
        # print("Direccionario solo con los dias!!!!!!!!!!!!!!!!!!!!!!!!!!!!!: ", diccionario)
        for d in diccionario:
                for e in entradas:
                        if d == e.timestamp.date():
                                diccionario[d] = diccionario[d] +1
                                # print(diccionario[d])
        clave_dicc = diccionario.keys()
        # print("indice o fecha: ", clave_dicc)
        valor_dicc = diccionario.values()
        # print("valores o cantidad de dias ", valor_dicc)
        cantidad_datos_dicc = diccionario.items()

    data = {
            'dias_total':dias_total,
            'clave_dicc': clave_dicc,
            'valor_dicc': valor_dicc,
            'cantidad_datos_dicc' : cantidad_datos_dicc,
            }
    return render(request, 'admin_grafico/graficos_linea/graficos_audios_admin.html',data)

#grafico cantidad intensidad
def grafico_intensidad_admin(request):

    oportunidades = Intensidad.objects.all()

    dias_total = []
    for e in oportunidades:
        dias_total.append(e.timestamp.date())
        # print(dias_total)
        # print("list original", dias_total)
        convert_list_to_set = set(dias_total)
        # print("Set is: ",convert_list_to_set)
        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)
        dias_total = list(convert_list_to_set)
        # print("Removed duplicates from original list: ",dias_total)
        dias_total.sort()
        #crear diccionario para grafico
        diccionario = {}
        for i in dias_total:
                        diccionario[i] = 0
        # print("Direccionario solo con los dias!!!!!!!!!!!!!!!!!!!!!!!!!!!!!: ", diccionario)
        for d in diccionario:
                for e in oportunidades:
                        if d == e.timestamp.date():
                                diccionario[d] = diccionario[d] +1
                                # print(diccionario[d])
        clave_dicc = diccionario.keys()
        # print("indice o fecha: ", clave_dicc)
        valor_dicc = diccionario.values()
        # print("valores o cantidad de dias ", valor_dicc)
        cantidad_datos_dicc = diccionario.items()

    data = {
            'dias_total':dias_total,
            'clave_dicc': clave_dicc,
            'valor_dicc': valor_dicc,
            'cantidad_datos_dicc' : cantidad_datos_dicc,
            }
    return render(request, 'admin_grafico/graficos_linea/graficos_intensidad_admin.html',data)

#grafico cantidad vocalizaciones
def grafico_vocalizacion_admin(request):


    entradas = Audio.objects.all()
    salidas = Vocalizacion.objects.all()
    oportunidades = Intensidad.objects.all()

    dias_total = []
    for e in salidas:
        dias_total.append(e.timestamp.date())
        # print(dias_total)
        # print("list original", dias_total)
        convert_list_to_set = set(dias_total)
        # print("Set is: ",convert_list_to_set)
        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)
        dias_total = list(convert_list_to_set)
        # print("Removed duplicates from original list: ",dias_total)
        dias_total.sort()
        #crear diccionario para grafico
        diccionario = {}
        for i in dias_total:
                        diccionario[i] = 0
        # print("Direccionario solo con los dias!!!!!!!!!!!!!!!!!!!!!!!!!!!!!: ", diccionario)
        for d in diccionario:
                for e in salidas:
                        if d == e.timestamp.date():
                                diccionario[d] = diccionario[d] +1
                                # print(diccionario[d])
        clave_dicc = diccionario.keys()
        # print("indice o fecha: ", clave_dicc)
        valor_dicc = diccionario.values()
        # print("valores o cantidad de dias ", valor_dicc)
        cantidad_datos_dicc = diccionario.items()

    data = {
            'dias_total':dias_total,
            'clave_dicc': clave_dicc,
            'valor_dicc': valor_dicc,
            'cantidad_datos_dicc' : cantidad_datos_dicc,
            }
    return render(request, 'admin_grafico/graficos_linea/graficos_vocalizacion_admin.html',data)

#grafico cantidad usuarios
def grafico_cantidad_usuarios(request):
    pacientes=Paciente.objects.all().count()
    neurologos=Neurologo.objects.all().count()    
    enfermeras=Enfermera.objects.all().count() 
    fonoaudiologos=Fonoaudiologo.objects.all().count()
    data={
        'pacientes':pacientes,
        'neurologos':neurologos,
        'enfermeras':enfermeras,
        'fonoaudiologos':fonoaudiologos
        
    }
    return render(request, 'admin_grafico/grafico_cantidad_usuarios/grafico_usuarios.html',data)

#grafico cantidad de pacientes por comuna
def grafico_pacientes_comuna(request):  
    # def color_aleatorio():
    #     r, g, b = random.randint(0,255), random.randint(0,255), random.randint(0,255)
    #     return "#{:02x}{:02x}{:02x}".format(r,g,b)
    # mil_colores = [color_aleatorio() for _ in range(350)]
    # print(mil_colores)
    salidas = Paciente.objects.all()
    dias_total = []
    for e in salidas:
        dias_total.append(e.comuna)
        # print(dias_total)
        # print("list original", dias_total)
        convert_list_to_set = set(dias_total)
        # print("Set is: ",convert_list_to_set)
        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)
        dias_total = list(convert_list_to_set)
        # print("Removed duplicates from original list: ",dias_total)
        #crear diccionario para grafico
        diccionario = {}
        for i in dias_total:
                        diccionario[i] = 0
        # print("Direccionario solo con los dias!!!!!!!!!!!!!!!!!!!!!!!!!!!!!: ", diccionario)
        for d in diccionario:
                for e in salidas:
                        if d == e.comuna:
                                diccionario[d] = diccionario[d] +1
                                # print(diccionario[d])
        clave_dicc = diccionario.keys()
        # print("indice o fecha: ", clave_dicc)
        valor_dicc = diccionario.values()
        # print("valores o cantidad de dias ", valor_dicc)
        cantidad_datos_dicc = diccionario.items()
    data = {
            'dias_total':dias_total,
            'clave_dicc': clave_dicc,
            'valor_dicc': valor_dicc,
            'cantidad_datos_dicc' : cantidad_datos_dicc,
            # 'mil_colores' : mil_colores,
            }
    return render(request, 'admin_grafico/graficos_comuna/grafico_pacientes_comuna.html',data)

#grafico cantidad de neurologo por comuna
def grafico_neurologo_comuna(request):  
    # def color_aleatorio():
    #     r, g, b = random.randint(0,255), random.randint(0,255), random.randint(0,255)
    #     return "#{:02x}{:02x}{:02x}".format(r,g,b)
    # mil_colores = [color_aleatorio() for _ in range(350)]
    # print(mil_colores)
    salidas = Neurologo.objects.all()
    dias_total = []
    for e in salidas:
        dias_total.append(e.comuna)
        # print(dias_total)
        # print("list original", dias_total)
        convert_list_to_set = set(dias_total)
        # print("Set is: ",convert_list_to_set)
        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)
        dias_total = list(convert_list_to_set)
        # print("Removed duplicates from original list: ",dias_total)
        #crear diccionario para grafico
        diccionario = {}
        for i in dias_total:
                        diccionario[i] = 0
        # print("Direccionario solo con los dias!!!!!!!!!!!!!!!!!!!!!!!!!!!!!: ", diccionario)
        for d in diccionario:
                for e in salidas:
                        if d == e.comuna:
                                diccionario[d] = diccionario[d] +1
                                # print(diccionario[d])
        clave_dicc = diccionario.keys()
        # print("indice o fecha: ", clave_dicc)
        valor_dicc = diccionario.values()
        # print("valores o cantidad de dias ", valor_dicc)
        cantidad_datos_dicc = diccionario.items()
    data = {
            'dias_total':dias_total,
            'clave_dicc': clave_dicc,
            'valor_dicc': valor_dicc,
            'cantidad_datos_dicc' : cantidad_datos_dicc,
            # 'mil_colores' : mil_colores,
            }
    return render(request, 'admin_grafico/graficos_comuna/grafico_neurologo_comuna.html',data)

#grafico cantidad de fonoaudiologo por comuna
def grafico_fonoaudiologo_comuna(request):  
    # def color_aleatorio():
    #     r, g, b = random.randint(0,255), random.randint(0,255), random.randint(0,255)
    #     return "#{:02x}{:02x}{:02x}".format(r,g,b)
    # mil_colores = [color_aleatorio() for _ in range(350)]
    # print(mil_colores)
    salidas = Fonoaudiologo.objects.all()
    dias_total = []
    for e in salidas:
        dias_total.append(e.comuna)
        # print(dias_total)
        # print("list original", dias_total)
        convert_list_to_set = set(dias_total)
        # print("Set is: ",convert_list_to_set)
        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)
        dias_total = list(convert_list_to_set)
        # print("Removed duplicates from original list: ",dias_total)
        #crear diccionario para grafico
        diccionario = {}
        for i in dias_total:
                        diccionario[i] = 0
        # print("Direccionario solo con los dias!!!!!!!!!!!!!!!!!!!!!!!!!!!!!: ", diccionario)
        for d in diccionario:
                for e in salidas:
                        if d == e.comuna:
                                diccionario[d] = diccionario[d] +1
                                # print(diccionario[d])
        clave_dicc = diccionario.keys()
        # print("indice o fecha: ", clave_dicc)
        valor_dicc = diccionario.values()
        # print("valores o cantidad de dias ", valor_dicc)
        cantidad_datos_dicc = diccionario.items()
    data = {
            'dias_total':dias_total,
            'clave_dicc': clave_dicc,
            'valor_dicc': valor_dicc,
            'cantidad_datos_dicc' : cantidad_datos_dicc,
            # 'mil_colores' : mil_colores,
            }
    return render(request, 'admin_grafico/graficos_comuna/grafico_fonoaudiologo_comuna.html',data)

#grafico cantidad de enfermeras por comuna
def grafico_enfermeras_comuna(request):  
    # def color_aleatorio():
    #     r, g, b = random.randint(0,255), random.randint(0,255), random.randint(0,255)
    #     return "#{:02x}{:02x}{:02x}".format(r,g,b)
    # mil_colores = [color_aleatorio() for _ in range(350)]
    # print(mil_colores)
    salidas = Enfermera.objects.all()
    dias_total = []
    for e in salidas:
        dias_total.append(e.comuna)
        # print(dias_total)
        # print("list original", dias_total)
        convert_list_to_set = set(dias_total)
        # print("Set is: ",convert_list_to_set)
        new_list = list(convert_list_to_set)
        #print("Resultant List is: ",new_list)
        dias_total = list(convert_list_to_set)
        # print("Removed duplicates from original list: ",dias_total)
        #crear diccionario para grafico
        diccionario = {}
        for i in dias_total:
                        diccionario[i] = 0
        # print("Direccionario solo con los dias!!!!!!!!!!!!!!!!!!!!!!!!!!!!!: ", diccionario)
        for d in diccionario:
                for e in salidas:
                        if d == e.comuna:
                                diccionario[d] = diccionario[d] +1
                                # print(diccionario[d])
        clave_dicc = diccionario.keys()
        # print("indice o fecha: ", clave_dicc)
        valor_dicc = diccionario.values()
        # print("valores o cantidad de dias ", valor_dicc)
        cantidad_datos_dicc = diccionario.items()
    data = {
            'dias_total':dias_total,
            'clave_dicc': clave_dicc,
            'valor_dicc': valor_dicc,
            'cantidad_datos_dicc' : cantidad_datos_dicc,
            # 'mil_colores' : mil_colores,
            }
    return render(request, 'admin_grafico/graficos_comuna/grafico_enfermeras_comuna.html',data)

#documentos para autorizar desde pacientes
def documento(request,username_paciente_id):
    documento=Documento.objects.all()
    paciente_documento=Paciente_Documento.objects.all().filter(username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    current_user =get_object_or_404(Paciente, pk=request.user.pk)
    if request.method == 'POST':
        form = FormDocumentos(request.POST)
        if form.is_valid():
            estado_emocion = form.save(commit=False)
            estado_emocion.username_paciente = current_user
            estado_emocion.save()
        return redirect('documento', current_user.username_paciente_id )
    else:
        form=FormDocumentos()
    return render(request, 'pacientes/documento.html',{'documento':documento,'pacientes':pacientes,'paciente_documento':paciente_documento,'form':form})

#listado de los documentos desde admin
def documento_admin(request):
    documento=Documento.objects.all()
    if request.method == 'POST':
        form = FormDocumentoAdmin(request.POST)
        if form.is_valid():
            estado_emocion = form.save(commit=False)
            estado_emocion.save()
        return redirect('documento_admin')
    else:
        form=FormDocumentoAdmin()
    return render(request, 'admin_documento/admin_documento.html',{'documento':documento,'form':form})

#listado de los preregistros
def preregistro_admin(request):
    preregistros=Preregistro.objects.all()
    par=Enfermera_neurologo.objects.all()
    return render(request, 'admin_paciente/preregistros.html',{'preregistros':preregistros,'par':par})



















#crear bitacora 
def agregar_bitacora(request,username_paciente_id):
    current_user =get_object_or_404(Enfermera, pk=request.user.pk)
    current_pac = get_object_or_404(Paciente, username_paciente_id=username_paciente_id)
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)


    if request.method == 'POST':
        form = FormBitacora(request.POST)
        if form.is_valid():
            estado_emocion = form.save(commit=False)
            estado_emocion.username_paciente = current_pac
            estado_emocion.username_enfermera = current_user
            estado_emocion.save()
        return redirect('enfermera_info_paciente', current_pac.username_paciente_id )
    else:
        form=FormBitacora()
    return render(request, 'enfermera/form_bitacora.html', {'form':form,'pacientes':pacientes})





def f0(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/f0.html',{'pacientes':pacientes})


def f0dev(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/f0dev.html',{'pacientes':pacientes})

def localabsolutejitter(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/localabsolutejitter.html',{'pacientes':pacientes})


def rapjitter(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/rapjitter.html',{'pacientes':pacientes})


def ppq5jitter(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/ppq5jitter.html',{'pacientes':pacientes})


def ddpjitter(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/ddpjitter.html',{'pacientes':pacientes})


def localshimer(request,username_paciente_id):

    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/localshimer.html',{'pacientes':pacientes})



def localdbshimer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/localdbshimer.html',{'pacientes':pacientes})



def apq3shimmer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/apq3shimmer.html',{'pacientes':pacientes})




def aqpq5shimer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/aqpq5shimer.html',{'pacientes':pacientes})



def apq11shimmer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/apq11shimmer.html',{'pacientes':pacientes})



def ddashimmer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/ddashimmer.html',{'pacientes':pacientes})



def f1(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/f1.html',{'pacientes':pacientes})


def f2(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/f2.html',{'pacientes':pacientes})




def f3(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/f3.html',{'pacientes':pacientes})



def f4(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/f4.html',{'pacientes':pacientes})

def graficos(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'fonoaudiologo/graficos/graficos.html',{'pacientes':pacientes})







def vista_f0(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/f0.html',{'pacientes':pacientes})


def vista_f0dev(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/f0dev.html',{'pacientes':pacientes})

def vista_localabsolutejitter(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/localabsolutejitter.html',{'pacientes':pacientes})


def vista_rapjitter(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/rapjitter.html',{'pacientes':pacientes})


def vista_ppq5jitter(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/ppq5jitter.html',{'pacientes':pacientes})


def vista_ddpjitter(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/ddpjitter.html',{'pacientes':pacientes})


def vista_localshimer(request,username_paciente_id):

    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/localshimer.html',{'pacientes':pacientes})



def vista_localdbshimer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/localdbshimer.html',{'pacientes':pacientes})



def vista_apq3shimmer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/apq3shimmer.html',{'pacientes':pacientes})




def vista_aqpq5shimer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/aqpq5shimer.html',{'pacientes':pacientes})



def vista_apq11shimmer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/apq11shimmer.html',{'pacientes':pacientes})



def vista_ddashimmer(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/ddashimmer.html',{'pacientes':pacientes})



def vista_f1(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/f1.html',{'pacientes':pacientes})


def vista_f2(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/f2.html',{'pacientes':pacientes})




def vista_f3(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/f3.html',{'pacientes':pacientes})



def vista_f4(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/f4.html',{'pacientes':pacientes})

def vista_graficos(request,username_paciente_id):
    pacientes=Paciente.objects.all().filter(username_paciente_id=username_paciente_id)
    return render(request, 'admin_paciente/vistas_fonoaudiologo/graficos/graficos.html',{'pacientes':pacientes})



#-------------------------------------------------------- Pre registros -----------------------------------------------------------------------------------------

#preregistros desde enfermera
def preregistros(request):
    print(request.user.tipo_usuario)
    if request.user.tipo_usuario.nombre_tipo_usuario == 'adminEnfermera':
        preregistros = Preregistro.objects.all()
    else:
        current_user=get_object_or_404(Enfermera_neurologo, username_enfermera=request.user.id)
        preregistros=Preregistro.objects.all().filter(neurologo=current_user.username_neurologo)

    return render(request, 'enfermera/preregistros.html',{'preregistros':preregistros})


#registro del usuario preregistrado
def registro_usuario(request,id):
    preregistros=Preregistro.objects.all().filter(id=id)

    data= {
        'form':PreregistroUsuarioForm(),
        'preregistros' :preregistros
    }
    if request.method == 'POST':
        formulario=PreregistroUsuarioForm(data=request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect("registro_paciente",id)
        data['form']=formulario
    return render(request,'enfermera/registro_usuario_paciente.html',data)



#registro del paciente preregistrado
class Registro_pacienteView(View):
    usuario =Usuario.objects.filter(tipo_usuario=1)
    comunas =Comuna.objects.all()
    diabete=Diabetes.objects.all()
    hipertensio=Hipertension.objects.all()
    template_name='enfermera/registro_paciente.html'

    def get(self,request,id):
        preregistros=Preregistro.objects.all().filter(id=id)
        total_pacientes=Paciente.objects.count()
        id_paciente = total_pacientes+1
        data={
            'usuario':self.usuario,
            'comunas':self.comunas,
            'diabete':self.diabete,
            'hipertensio':self.hipertensio,
            'preregistros':preregistros,
            'id_paciente':id_paciente,
        }
        return render(request,self.template_name,data)

    def post(self,request,id):
        id_paciente =request.POST['id']
        username_paciente=request.POST['username_paciente']
        rut_paciente=request.POST['rut_paciente']
        nombre_paciente=request.POST['nombre_paciente']
        apellido_paciente=request.POST['apellido_paciente']
        diabetes=request.POST['diabetes']
        hipertension=request.POST['hipertension']
        direccion_paciente=request.POST['direccion_paciente']
        comuna=request.POST['comuna']
        email_paciente=request.POST['email_paciente']
        telefono_paciente=request.POST['telefono_paciente']
        whatsaap_paciente=request.POST['whatsaap_paciente']
        celular_paciente=request.POST['celular_paciente']
        telegram_paciente=request.POST['telegram_paciente']
        new_paciente=Paciente.objects.create(id_paciente=id_paciente,
                                            username_paciente_id=username_paciente,
                                            rut_paciente=rut_paciente,
                                            nombre_paciente=nombre_paciente,
                                            apellido_paciente=apellido_paciente,
                                            diabetes_id=diabetes,
                                            hipertension_id=hipertension,
                                            direccion_paciente=direccion_paciente,
                                            comuna_id=comuna,
                                            email_paciente=email_paciente,
                                            telefono_paciente=telefono_paciente,
                                            whatsaap_paciente=whatsaap_paciente,
                                            celular_paciente=celular_paciente,
                                            telegram_paciente=telegram_paciente
                                            )
        new_paciente.save()
        return redirect("enf_pac", id)


#modificar audios
def modificar_audio(request, id_audio):
    audio = get_object_or_404(Audio, id=id_audio)
    get_audio = Audio.objects.get(id = id_audio)
    id_paciente = get_audio.username_paciente_id
    audios =Audio.objects.all().filter(id=id_audio)
    
    data = {
        'form': AudioForm(instance=audio),
        'audios':audios
    }
    if request.method == 'POST':
        formulario = AudioForm(data=request.POST, instance=audio, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            return redirect(f'/vista_audio_paciente/{id_paciente}')
        else:
            data["form"] = formulario
    return render(request, 'admin_paciente/vistas_fonoaudiologo/editar_audio.html', data)

#modificar variables de audio
def modificar_audio_fonoaudiologo(request, id_audio):
    audio = get_object_or_404(Audio, id=id_audio)
    get_audio = Audio.objects.get(id = id_audio)
    id_paciente = get_audio.username_paciente_id
    audios =Audio.objects.all().filter(id=id_audio)
    
    data = {
        'form': AudioForm(instance=audio),
        'audios':audios
    }
    if request.method == 'POST':
        formulario = AudioForm(data=request.POST, instance=audio, files=request.FILES)
        if formulario.is_valid():
            formulario.save()
            return redirect(f'/audio_paciente/{id_paciente}')
        else:
            data["form"] = formulario
    return render(request, 'fonoaudiologo/editar_audio.html', data)


#home para enviar correo
class HomeCorreosView(View):
    template_name = "admin_correo/home_envio_correo.html"
    
    def get(self, request, *args, **kwargs):
       
        return render(request, self.template_name)

#escribir correo para el usuario seleccionado
class EnviarCorreoView(View): 
    template_name = "admin_correo/form_enviar_correos.html" 
    registro_correos = RegistroCorreos.objects.all() 
    #correos = ["fabian.salazar.se12@gmail.com", "vozparkinsonchile@gmail.com"] 
     
    def get(self, request, usuario): 
        data = { 
            "usuario":usuario, 
            "registro_correos":self.registro_correos, 
        } 
 
        return render(request, self.template_name, data) 
 
    def post(self, request, usuario): 
        ##enviar correo a fonoaudiologos 
        if usuario == "1": 
            correos = Fonoaudiologo.objects.all()     
            for c in correos: 
                correos = c.email_fonoaudiologo 
                subject = request.POST['asunto'] 
                message = request.POST['mensaje'] 
                send_mail(subject, message, 'energyfabian12@gmail.com',[correos], fail_silently=False) 
                registro_mensaje = RegistroCorreos.objects.create(asunto = subject, mensaje = message, dirigido = "fonoaudiologos") 
                registro_mensaje.save() 
            messages.success(request, "Correo enviado")
            return redirect('home_enviar_correo') 
 
            
        #enviar correo e enfermera 
        elif usuario == "2": 
            correos = Enfermera.objects.all()     
            for c in correos: 
                correos = c.email_enfermera 
                subject = request.POST['asunto'] 
                message = request.POST['mensaje'] 
                send_mail(subject, message, 'energyfabian12@gmail.com',[correos], fail_silently=False) 
                registro_mensaje = RegistroCorreos.objects.create(asunto = subject, mensaje = message, dirigido = "enfermeras") 
                registro_mensaje.save() 
            messages.success(request, "Correo enviado")
 
            return redirect('home_enviar_correo') 
        #enviar correo e paciente 
        elif usuario == "3": 
            correos = Paciente.objects.all()     
            for c in correos: 
                correos = c.email_paciente 
                subject = request.POST['asunto'] 
                message = request.POST['mensaje'] 
                send_mail(subject, message, 'energyfabian12@gmail.com',[correos], fail_silently=False) 
                registro_mensaje = RegistroCorreos.objects.create(asunto = subject, mensaje = message, dirigido = "pacientes") 
                registro_mensaje.save() 
            messages.success(request, "Correo enviado")
            return redirect('home_enviar_correo')  
        #enviar correo a familiares 
        elif usuario == "4": 
            correos = Familiar.objects.all()     
            for c in correos: 
                correos = c.email_familiar 
                subject = request.POST['asunto'] 
                message = request.POST['mensaje'] 
                send_mail(subject, message, 'energyfabian12@gmail.com',[correos], fail_silently=False)  
                registro_mensaje = RegistroCorreos.objects.create(asunto = subject, mensaje = message, dirigido = "familiares") 
                registro_mensaje.save()   
            messages.success(request, "Correo enviado")
            return redirect('home_enviar_correo')  
        else: 
            return HttpResponse('Correo no enviado')
